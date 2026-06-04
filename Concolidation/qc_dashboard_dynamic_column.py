import os
import openpyxl
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

# Import structural theme functions and constants from our design module
import Theme_engine

def find_actual_template_width(filepath):
    """
    Scans the template's first two rows to find the absolute last column index
    that contains a real, non-empty value, stripping out ghost cells or empty spaces.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Scan up to 250 columns out to find where the actual header values drop off
        check_limit = ws.max_column if (ws.max_column and ws.max_column > 20) else 250
        
        row1_vals = [ws.cell(row=1, column=col).value for col in range(1, check_limit + 1)]
        row2_vals = [ws.cell(row=2, column=col).value for col in range(1, check_limit + 1)]
        wb.close()
        
        # Strip trailing None fields from the end of the arrays
        while row1_vals and row1_vals[-1] is None and row2_vals[-1] is None:
            row1_vals.pop()
            row2_vals.pop()
            
        actual_width = max(len(row1_vals), len(row2_vals))
        return actual_width if actual_width > 0 else 20
    except Exception:
        return 50  # Safe default fallback if file reading error occurs

class QCDashboard:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Pipeline Header QC System - Adaptive Scale Engine")
        
        # --- DYNAMIC MONITOR CENTERING LOGIC ---
        window_width = 1200
        window_height = 800
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        center_x = int((screen_width / 2) - (window_width / 2))
        center_y = int((screen_height / 2) - (window_height / 2))
        
        self.root.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")
        self.root.resizable(False, False) # Keep layout structured and fixed
        
        self.template_path = ""
        self.target_dir = ""
        
        # Initialize our modularized theme layouts across elements
        self.style = Theme_engine.apply_global_theme(self.root)
        
        # Build out Layout structures
        self.build_ui()
        
    def build_ui(self):
        # 1. Top Header Banner (Packed to root)
        header_frame = ttk.Frame(self.root, style="Header.TFrame", padding=15)
        header_frame.pack(fill="x", side="top")
        
        header_lbl = ttk.Label(header_frame, text="📊 DATA PIPELINE COMPLIANCE DASHBOARD (ADAPTIVE)", style="HeaderLabel.TLabel")
        header_lbl.pack(anchor="w")

        # 2. Main Workspace Panel Settings (Packed to root)
        workspace = ttk.Frame(self.root, padding=20, style="TFrame")
        workspace.pack(fill="x", side="top")
        
        # Configure grid column weights so the middle column stretches automatically
        workspace.columnconfigure(1, weight=1)
        
        # Row 1: Template Selection Layout
        lbl_temp = ttk.Label(workspace, text="Benchmark Target Layout:", style="Standard.TLabel")
        lbl_temp.grid(row=0, column=0, sticky="w", pady=8, padx=5)
        
        self.lbl_temp_path = ttk.Label(workspace, text=" No file selected...", style="Path.TLabel", width=100, anchor="w")
        self.lbl_temp_path.grid(row=0, column=1, sticky="ew", padx=10, pady=8, ipady=4)
        
        btn_browse_temp = ttk.Button(workspace, text="Browse Template File", style="Action.TButton", command=self.select_template)
        btn_browse_temp.grid(row=0, column=2, padx=5, pady=8)
        
        # Row 2: Target Audit Folder Setup
        lbl_folder = ttk.Label(workspace, text="QC Audit Backup Directory:", style="Standard.TLabel")
        lbl_folder.grid(row=1, column=0, sticky="w", pady=8, padx=5)
        
        self.lbl_folder_path = ttk.Label(workspace, text=" No folder selected...", style="Path.TLabel", width=100, anchor="w")
        self.lbl_folder_path.grid(row=1, column=1, sticky="ew", padx=10, pady=8, ipady=4)
        
        btn_browse_folder = ttk.Button(workspace, text="Browse Audit Folder", style="Action.TButton", command=self.select_folder)
        btn_browse_folder.grid(row=1, column=2, padx=5, pady=8)
        
        # Row 3: Evaluation Runtime Processing Action Call Block
        self.btn_run = ttk.Button(workspace, text="⚡ RUN COMBINED MATRIX AUDIT", style="Run.TButton", command=self.execute_qc, state="disabled")
        self.btn_run.grid(row=2, column=0, columnspan=3, pady=20, ipady=5)
        
        # 3. Integrated Shell Terminal Logging Space Frame (Packed to root)
        console_frame = ttk.Frame(self.root, padding=15)
        console_frame.pack(fill="both", expand=True, side="top")
        
        console_lbl = ttk.Label(console_frame, text="Live Validation Logs Output (Expanded Terminal View):", style="Standard.TLabel")
        console_lbl.pack(anchor="w", pady=5)
        
        # Setup terminal text log element
        self.log_box = tk.Text(console_frame)
        scrollbar = ttk.Scrollbar(console_frame, orient="vertical", command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        self.log_box.pack(side="left", fill="both", expand=True)
        
        # Apply custom design properties to non-ttk elements via our theme helper module
        Theme_engine.style_console_widget(self.log_box)

    def log(self, message, tag=None):
        self.log_box.insert(tk.END, message + "\n", tag)
        self.log_box.see(tk.END)

    def select_template(self):
        file = filedialog.askopenfilename(title="Select Benchmark File", filetypes=[("Excel Workspace", "*.xlsx *.xls")])
        if file:
            self.template_path = file
            self.lbl_temp_path.config(text=f"  {file}")
            self.check_ready_state()

    def select_folder(self):
        folder = filedialog.askdirectory(title="Select Folder to Audit")
        if folder:
            self.target_dir = folder
            self.lbl_folder_path.config(text=f"  {folder}")
            self.check_ready_state()

    def check_ready_state(self):
        if self.template_path and self.target_dir:
            self.btn_run.state(["!disabled"])

    # ==============================================================================
    # HEADER VALIDATION PIPELINE EXECUTION
    # ==============================================================================
    def get_header_structure(self, filepath, max_cols_limit):
        """
        Extracts a nested list structure representing Row 1 and Row 2 headers.
        Enforces a strict evaluation column width defined by the dynamic benchmark rule.
        """
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            
            # Read exactly up to our calculated limit boundary rule to prevent trailing grid noise
            row1_vals = [ws.cell(row=1, column=col).value for col in range(1, max_cols_limit + 1)]
            row2_vals = [ws.cell(row=2, column=col).value for col in range(1, max_cols_limit + 1)]
            wb.close()
            
            # Clean trailing blank buffer cells for local evaluation
            while row1_vals and row1_vals[-1] is None and row2_vals[-1] is None:
                row1_vals.pop()
                row2_vals.pop()
                
            structure, current_parent, current_children = [], None, []
            for r1, r2 in zip(row1_vals, row2_vals):
                if r1 is not None:
                    if current_parent is not None:
                        structure.append((current_parent, current_children))
                    current_parent = r1
                    current_children = [r2] if r2 is not None else []
                else:
                    if r2 is not None:
                        current_children.append(r2)
            if current_parent is not None:
                structure.append((current_parent, current_children))
            return structure
        except Exception as e:
            return None

    def diagnose_structural_differences(self, template_struct, target_struct):
        anomalies = []
        template_map = {g: s for g, s in template_struct}
        target_map = {g: s for g, s in target_struct}
        
        for group in template_map:
            if group not in target_map:
                anomalies.append(f"Missing Section Group: '{group}'")
        for group in target_map:
            if group not in template_map:
                anomalies.append(f"Unexpected Group Found: '{group}'")
                
        for group in template_map:
            if group in target_map:
                t_cols, tgt_cols = template_map[group], target_map[group]
                if t_cols != tgt_cols:
                    missing_cols = [c for c in t_cols if c not in tgt_cols]
                    extra_cols = [c for c in tgt_cols if c not in t_cols]
                    
                    if missing_cols:
                        anomalies.append(f"Group '{group}': Missing column(s) {missing_cols}")
                    if extra_cols:
                        anomalies.append(f"Group '{group}': Extra unexpected column(s) {extra_cols}")
                    
                    if len(t_cols) == len(tgt_cols):
                        for idx, (t_col, tgt_col) in enumerate(zip(t_cols, tgt_cols)):
                            if t_col != tgt_col:
                                anomalies.append(f"Group '{group}': Typo detected! Expected '{t_col}' but found '{tgt_col}' at col {idx+1}")
        return anomalies

    def execute_qc(self):
        self.log_box.delete("1.0", tk.END)
        self.log("="*110, "INFO")
        self.log(" STARTING ADAPTIVE GRID SCALE MATRIX AUDIT RUN", "INFO")
        self.log("="*110, "INFO")
        
        # 1. Print Selected Workspace File Profiles right at the top
        self.log(f"[BLUEPRINT FILE] : {os.path.basename(self.template_path)}")
        self.log(f"[AUDIT TARGET]   : {self.target_dir}")
        self.log("-"*110, "INFO")
        
        # 2. Dynamically figure out how wide the template *actually* is (stripping trailing whitespace)
        template_actual_width = find_actual_template_width(self.template_path)
        
        # 3. Add custom margin of +5 safety buffer spaces
        evaluation_max_cols = template_actual_width + 5
        
        self.log(f"[CONFIG] Base template actual header size detected: {template_actual_width} columns.")
        self.log(f"[CONFIG] Setting target scanner dynamic boundary window to: {evaluation_max_cols} columns (+5 safety margin).")
        self.log("-"*110, "INFO")
        
        # 4. Pass our dynamic evaluation range into our reader module
        template_struct = self.get_header_structure(self.template_path, evaluation_max_cols)
        if not template_struct:
            self.log("[CRITICAL ERROR] Failed parsing Benchmark template schema.", "FAIL")
            return
            
        excel_files = [f for f in os.listdir(self.target_dir) if f.endswith(('.xlsx', '.xls'))]
        if not excel_files:
            self.log("[-] Aborted: No valid target excel sheets inside directory.", "FAIL")
            return
            
        pass_count, fail_count = 0, 0
        
        for file in sorted(excel_files):
            file_path = os.path.join(self.target_dir, file)
            # 5. Force target files to scan using the exact same calculated layout rule
            target_struct = self.get_header_structure(file_path, evaluation_max_cols)
            
            if target_struct is None:
                self.log(f"❌ {file} | File unreadable / corrupted file structure.", "FAIL")
                fail_count += 1
                continue
                
            errors = self.diagnose_structural_differences(template_struct, target_struct)
            
            if not errors:
                self.log(f"✔ {file} -> Perfect Match.", "PASS")
                pass_count += 1
            else:
                self.log(f"❌ {file} -> Anomalies Found:", "FAIL")
                for err in errors:
                    self.log(f"    ↳ ERROR: {err}", "FAIL")
                fail_count += 1
                
        self.log("\n" + "="*110, "INFO")
        self.log(f" AUDIT DONE | Total: {len(excel_files)} | Passed: {pass_count} | Failed: {fail_count}", "INFO")
        self.log("="*110, "INFO")
        messagebox.showinfo("Finished", "Adaptive batch validation finished successfully!")

if __name__ == "__main__":
    main_window = tk.Tk()
    app = QCDashboard(main_window)
    main_window.mainloop()